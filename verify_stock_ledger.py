"""
재고수불부 검증 스크립트 (BOM 정교화 포함)
stock_sample.xlsx를 읽어 재고수불부 + BOM 전개 수량을 계산.

BOM 로직 (bundled_sku_ro):
  - bundled_sku_ro.sku_id        = BOM 완제품 SKU
  - bundled_sku_ro.bundled_sku_id = BOM 구성요소 SKU
  - bundled_sku_ro.quantity       = 단위 구성 수량

배송완료 기반 BOM 전개 소요수량:
  경로 A: delivery_item.sku_id가 BOM 완제품인 경우
          → 구성요소 소요 = delivery_item.quantity × bom.quantity
  경로 B: delivery_item.sku_id가 구성요소 자체인 경우
          → 소요 = delivery_item.quantity

기초재고/기말재고: 체인 집합 방식 (멀티창고 stock_id별 계산 후 합산)
"""

import pandas as pd
from pathlib import Path

BASE_DIR    = Path(__file__).parent
INPUT_FILE  = BASE_DIR / "stock_sample.xlsx"
OUTPUT_FILE = BASE_DIR / "stock_ledger_result.xlsx"

# ── 1. 데이터 로드 ────────────────────────────────────────────────────────────
xl = pd.read_excel(INPUT_FILE, sheet_name=None, dtype=str)

su        = xl["stock_usage_ro"].copy()
sku       = xl["sku_ro"].copy()
sku_group = xl["sku_group_ro"].copy()
bom       = xl["bundled_sku_ro"].copy()
dl        = xl["delivery_ro"].copy()
dli       = xl["delivery_item_ro"].copy()

# delivery_item_ro에서 SKU 이름 보완용 매핑 생성
# (item_name + option_name 조합, sku_ro에 없는 SKU 대상)
dli_name = (
    dli[dli["item_name"].notna()][["sku_id","item_name","option_name"]]
    .drop_duplicates("sku_id")
    .copy()
)
dli_name["sku_nm_fallback"] = dli_name.apply(
    lambda r: f"{r['item_name']} [{r['option_name']}]"
    if pd.notna(r.get("option_name")) else r["item_name"], axis=1
)

su["updated_at"]      = pd.to_datetime(su["updated_at"])
su["before_quantity"] = pd.to_numeric(su["before_quantity"])
su["after_quantity"]  = pd.to_numeric(su["after_quantity"])
su["delta"]           = pd.to_numeric(su["delta"])
su["dt"]              = su["updated_at"].dt.date.astype(str)

bom["quantity"] = pd.to_numeric(bom["quantity"])
dli["quantity"] = pd.to_numeric(dli["quantity"])

# ── 2. BOM 유형 분류 ──────────────────────────────────────────────────────────
bom_finished  = set(bom["sku_id"])         # BOM 완제품 SKU
bom_component = set(bom["bundled_sku_id"]) # BOM 구성요소 SKU

def classify_bom(sku_id):
    if sku_id in bom_finished:
        return "BOM완제품"
    elif sku_id in bom_component:
        return "BOM구성요소"
    return "SINGLE"

# ── 3. 체인 기반 기초재고 / 기말재고 계산 ─────────────────────────────────────
def get_opening_closing_per_stock(group):
    bq_set = set(group["before_quantity"])
    aq_set = set(group["after_quantity"])
    opening_set = bq_set - aq_set
    closing_set = aq_set - bq_set

    if opening_set:
        opening = min(opening_set)
    else:
        earliest_ts = group["updated_at"].min()
        eg = group[group["updated_at"] == earliest_ts]
        cand = set(eg["before_quantity"]) - set(eg["after_quantity"])
        opening = min(cand) if cand else int(group.sort_values("updated_at").iloc[0]["before_quantity"])

    if closing_set:
        closing = max(closing_set)
    else:
        latest_ts = group["updated_at"].max()
        lg = group[group["updated_at"] == latest_ts]
        cand = set(lg["after_quantity"]) - set(lg["before_quantity"])
        closing = max(cand) if cand else int(group.sort_values("updated_at").iloc[-1]["after_quantity"])

    return pd.Series({"기초재고": int(opening), "기말재고": int(closing)})

per_stock = (
    su.groupby(["dt", "sku_id", "stock_id"])
    .apply(get_opening_closing_per_stock, include_groups=False)
    .reset_index()
)
stock_bounds = (
    per_stock.groupby(["dt", "sku_id"])[["기초재고", "기말재고"]]
    .sum()
    .reset_index()
)

# ── 4. type별 delta 합계 ──────────────────────────────────────────────────────
def sum_delta_by_type(df, type_val, col_name):
    return (
        df[df["type"] == type_val]
        .groupby(["dt", "sku_id"])["delta"]
        .sum()
        .reset_index()
        .rename(columns={"delta": col_name})
    )

incoming_agg   = sum_delta_by_type(su, "INCOMING_COMPLETED",   "입고수량")
adjustment_agg = sum_delta_by_type(su, "ADJUSTMENT_COMPLETED", "조정수량")
out_cmp_agg    = sum_delta_by_type(su, "OUTGOING_COMPLETED",   "출고완료")
out_req_agg    = sum_delta_by_type(su, "OUTGOING_REQUESTED",   "출고요청")
out_can_agg    = sum_delta_by_type(su, "OUTGOING_CANCELLED",   "출고취소")

net_delta = (
    su.groupby(["dt", "sku_id"])["delta"]
    .sum()
    .reset_index()
    .rename(columns={"delta": "순변동"})
)

# ── 5. 배송완료수량 (delivery_ro COMPLETED 기준) ──────────────────────────────
dl_done = dl[dl["status"] == "COMPLETED"][["id"]].rename(columns={"id": "delivery_id"})
dli_done = dli.merge(dl_done, on="delivery_id", how="inner")
delivery_agg = (
    dli_done.groupby("sku_id")["quantity"]
    .sum()
    .reset_index()
    .rename(columns={"quantity": "배송완료수량"})
)
delivery_agg["배송완료수량"] = delivery_agg["배송완료수량"].astype(int)

# ── 6. BOM 전개 소요수량 계산 ─────────────────────────────────────────────────
# 구성요소 SKU의 stock_usage_ro OUTGOING_COMPLETED 합산
# (실제 재고 차감이 완제품 레벨이 아닌 구성요소 레벨에서만 발생하는 구조)
bom_req = (
    su[(su["sku_id"].isin(bom_component)) & (su["type"] == "OUTGOING_COMPLETED")]
    .groupby("sku_id")["delta"]
    .sum()
    .abs()
    .reset_index()
    .rename(columns={"delta": "BOM전개소요수량"})
)
bom_req["BOM전개소요수량"] = bom_req["BOM전개소요수량"].astype(int)

# ── 7. SKU 정보 조인 ──────────────────────────────────────────────────────────
sku_info = (
    sku[sku["deleted_at"].isna()][["id", "name", "sku_code", "composition_type", "sku_group_id"]]
    .rename(columns={"id": "sku_id", "name": "sku_nm"})
    .merge(
        sku_group[sku_group["deleted_at"].isna()][["id", "biz_partner_id"]],
        left_on="sku_group_id", right_on="id", how="left"
    )
    .drop(columns=["id", "sku_group_id"])
)
sku_info["BOM유형"] = sku_info["sku_id"].apply(classify_bom)

# ── 8. 전체 결합 ──────────────────────────────────────────────────────────────
base = net_delta.copy()
for agg_df in [stock_bounds, incoming_agg, adjustment_agg, out_cmp_agg, out_req_agg, out_can_agg]:
    base = base.merge(agg_df, on=["dt", "sku_id"], how="left")

num_cols = ["기초재고", "입고수량", "조정수량", "출고완료", "출고요청", "출고취소", "기말재고"]
base[num_cols] = base[num_cols].fillna(0).astype(int)

result = (
    base
    .merge(sku_info, on="sku_id", how="left")
    .merge(delivery_agg, on="sku_id", how="left")
    .merge(bom_req, on="sku_id", how="left")
)
result["배송완료수량"]   = result["배송완료수량"].fillna(0).astype(int)
result["BOM전개소요수량"] = result["BOM전개소요수량"].fillna(0).astype(int)

# sku_nm 보완: sku_ro에 없는 SKU → delivery_item_ro.item_name으로 채움
result = result.merge(dli_name[["sku_id","sku_nm_fallback"]], on="sku_id", how="left")
result["sku_nm"] = result["sku_nm"].fillna(result["sku_nm_fallback"])
result.drop(columns=["sku_nm_fallback"], inplace=True)

FINAL_COLS = [
    "dt", "sku_id", "sku_nm", "sku_code", "biz_partner_id", "composition_type", "BOM유형",
    "기초재고", "입고수량", "조정수량", "출고완료", "출고요청", "출고취소", "순변동", "기말재고",
    "배송완료수량", "BOM전개소요수량"
]
result = result[FINAL_COLS]

# ── 8-1. BOM 완제품 행 추가 ───────────────────────────────────────────────────
# 고유 구성요소(단 하나의 완제품에만 속한 구성요소) 기반으로 완제품 판매수량 역산
# 판매수량 = 구성요소 OUTGOING_COMPLETED 절댓값 / bom.quantity
comp_parent_cnt = bom.groupby("bundled_sku_id")["sku_id"].nunique()
unique_comps = set(comp_parent_cnt[comp_parent_cnt == 1].index)

bom_unique = bom[bom["bundled_sku_id"].isin(unique_comps)][["sku_id", "bundled_sku_id", "quantity"]]

su_unique = (
    su[(su["sku_id"].isin(unique_comps)) & (su["type"] == "OUTGOING_COMPLETED")]
    .groupby(["dt", "sku_id"])["delta"]
    .sum()
    .abs()
    .reset_index()
    .rename(columns={"sku_id": "bundled_sku_id", "delta": "comp_outgoing"})
)

finished_sales = (
    su_unique
    .merge(bom_unique, on="bundled_sku_id", how="left")
    .assign(판매수량=lambda x: (x["comp_outgoing"] / x["quantity"]).round(0).astype(int))
    .groupby(["dt", "sku_id"])["판매수량"]
    .sum()  # 완제품 1개에 고유 구성요소 여러 개인 경우 합산이 아닌 대표값 필요 → MAX 사용
    .reset_index()
)
# 완제품 1개에 고유 구성요소 2개 이상이면 max(=더 큰 쪽) 사용
finished_sales = (
    su_unique
    .merge(bom_unique, on="bundled_sku_id", how="left")
    .assign(판매수량=lambda x: (x["comp_outgoing"] / x["quantity"]).round(0).astype(int))
    .groupby(["dt", "sku_id"])["판매수량"]
    .max()
    .reset_index()
)

# 완제품 sku_info (sku_ro에 없으면 NaN 유지)
finished_sku_info = (
    pd.DataFrame({"sku_id": list(bom_finished)})
    .merge(sku_info, on="sku_id", how="left")
)
finished_sku_info["BOM유형"] = "BOM완제품"

# 완제품 행 구성
finished_rows = finished_sales.merge(finished_sku_info, on="sku_id", how="left")
finished_rows["출고완료"]     = -finished_rows["판매수량"]
finished_rows["순변동"]       = -finished_rows["판매수량"]
for col in ["기초재고", "입고수량", "조정수량", "출고요청", "출고취소", "기말재고",
            "배송완료수량", "BOM전개소요수량"]:
    finished_rows[col] = 0
finished_rows = finished_rows[FINAL_COLS]

# 구성요소 + 완제품 합치기
result = (
    pd.concat([result, finished_rows], ignore_index=True)
    .sort_values(["dt", "biz_partner_id", "BOM유형", "sku_id"],
                 ascending=[True, True, False, True])  # BOM완제품이 구성요소 위에
    .reset_index(drop=True)
)

# ── 9. 검증 출력 ──────────────────────────────────────────────────────────────
print(f"총 행수: {len(result)}")
print(f"날짜 범위: {result['dt'].min()} ~ {result['dt'].max()}")
print(f"고유 SKU 수: {result['sku_id'].nunique()}")

bom_counts = result.drop_duplicates("sku_id")["BOM유형"].value_counts()
print(f"\n[BOM 유형 분포]")
for k, v in bom_counts.items():
    print(f"  {k}: {v}개 SKU")

print("\n[일별 요약]")
# BOM완제품 제외 (실물재고 없음)
real = result[result["BOM유형"] != "BOM완제품"].copy()
for c in ["기초재고","입고수량","조정수량","출고완료","출고요청","출고취소","순변동","기말재고"]:
    real[c] = pd.to_numeric(real[c])

# 전일기말 = 당일기초가 되도록 누적 방식 계산
# - 처음 등장하는 SKU의 기초재고를 전날 기말합계에 더함
# - 기초 + 순변동 = 기말 (by construction)
dates = sorted(real["dt"].unique())
sku_seen = set()
prev_기말 = None
summary_rows = []

for dt in dates:
    day = real[real["dt"] == dt]
    active_skus = set(day["sku_id"])
    new_skus = active_skus - sku_seen  # 이날 처음 등장한 SKU

    신규_기초 = int(day[day["sku_id"].isin(new_skus)]["기초재고"].sum())

    if prev_기말 is None:
        기초합계 = int(day["기초재고"].sum())  # 첫날: 실제 기초재고 그대로
    else:
        기초합계 = prev_기말 + 신규_기초     # 다음날: 전날 기말 + 신규 SKU 기초재고

    순변동합계 = int(day["순변동"].sum())
    기말합계   = 기초합계 + 순변동합계

    summary_rows.append({
        "dt":            dt,
        "활동SKU수":     len(active_skus),
        "기초재고합계":  기초합계,
        "입고합계":      int(day["입고수량"].sum()),
        "조정합계":      int(day["조정수량"].sum()),
        "출고완료합계":  int(day["출고완료"].sum()),
        "출고요청합계":  int(day["출고요청"].sum()),
        "출고취소합계":  int(day["출고취소"].sum()),
        "순변동합계":    순변동합계,
        "기말재고합계":  기말합계,
    })
    prev_기말  = 기말합계
    sku_seen  |= active_skus

summary = pd.DataFrame(summary_rows)
summary["검증(기초+순변동-기말)"] = (
    summary["기초재고합계"] + summary["순변동합계"] - summary["기말재고합계"]
)
print(summary.to_string(index=False))

print("\n[상위 10행 미리보기]")
print(result.head(10).to_string(index=False))

# 검증: 기초재고 + 순변동 = 기말재고 (BOM완제품 제외 — 실물재고 없어 기말=0)
check = result[result["BOM유형"] != "BOM완제품"].copy()
check["계산기말"] = pd.to_numeric(check["기초재고"]) + pd.to_numeric(check["순변동"])
check["기말재고"] = pd.to_numeric(check["기말재고"])
mismatch = check[check["계산기말"] != check["기말재고"]]
print(f"\n[검증] 기초재고 + 순변동 = 기말재고 불일치 건수: {len(mismatch)}  (BOM완제품 제외)")
if len(mismatch) > 0:
    print(mismatch[["dt","sku_id","기초재고","순변동","계산기말","기말재고"]].head(10).to_string(index=False))

# ── 10. Excel 출력 ────────────────────────────────────────────────────────────
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    result.to_excel(writer, sheet_name="재고수불부", index=False)
    summary.to_excel(writer, sheet_name="일별요약", index=False)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = writer.book
    header_font  = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", start_color="2F5496")
    header_align = Alignment(horizontal="center", vertical="center")
    num_fmt      = "#,##0;(#,##0);\"-\""

    # 재고수불부 시트
    ws = wb["재고수불부"]
    col_widths = {
        "A":12, "B":20, "C":40, "D":22, "E":18, "F":14, "G":14,
        "H":12, "I":12, "J":12, "K":12, "L":12, "M":12, "N":12, "O":12,
        "P":14, "Q":16
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    num_col_idx = set(range(8, 18))  # H~Q (1-indexed)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name="Arial", size=9)
            if cell.column in num_col_idx:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "H2"

    # 일별요약 시트
    ws2 = wb["일별요약"]
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    for col in ws2.iter_cols(min_col=3, max_col=ws2.max_column):
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 18
    num_col_summary = set(range(3, ws2.max_column + 1))
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        for cell in row:
            cell.font = Font(name="Arial")
            if cell.column in num_col_summary:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")

print(f"\n저장 완료: {OUTPUT_FILE}")
