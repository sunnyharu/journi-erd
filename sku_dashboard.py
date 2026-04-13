"""
SKU 통합 현황 대시보드
─────────────────────────────────────────────────────────────────
재고 가용 현황 + 출고 이행률 + 기간 재고 변동 + 주문/배송을 SKU 기준으로 통합.

시트 구성:
  Sheet1. 전체현황   — 전체 SKU 통합 뷰
  Sheet2. 미처리출고  — 미처리수량 > 0 SKU (물류팀 확인용)
  Sheet3. 재고부족위험 — 가용재고 < 미처리수량 (MD/운영팀 긴급 확인용)
  Sheet4. 이행완료   — 이행률 100% SKU (정상 처리 확인용)
"""

import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

BASE_DIR    = Path(__file__).parent
INPUT_FILE  = BASE_DIR / "stock_sample.xlsx"
OUTPUT_FILE = BASE_DIR / "sku_dashboard.xlsx"

# ── 1. 데이터 로드 ────────────────────────────────────────────
xl      = pd.read_excel(INPUT_FILE, sheet_name=None, dtype=str)
sku     = xl["sku_ro"].copy()
sku_grp = xl["sku_group_ro"].copy()
stock   = xl["stock_ro"].copy()
og      = xl["outgoing_ro"].copy()
ogi     = xl["outgoing_item_ro"].copy()
su      = xl["stock_usage_ro"].copy()
dl      = xl["delivery_ro"].copy()
dli     = xl["delivery_item_ro"].copy()
ol      = xl["order_line_ro"].copy()
bom     = xl["bundled_sku_ro"].copy()

for col in ["physical_quantity","available_quantity","reserved_quantity"]:
    stock[col] = pd.to_numeric(stock[col])
ogi["quantity"]          = pd.to_numeric(ogi["quantity"])
ogi["outgoing_quantity"] = pd.to_numeric(ogi["outgoing_quantity"]).fillna(0)
su["delta"]              = pd.to_numeric(su["delta"])
dli["quantity"]          = pd.to_numeric(dli["quantity"])
ol["quantity"]           = pd.to_numeric(ol["quantity"])
ol["base_currency_amount"] = pd.to_numeric(ol["base_currency_amount"])
bom["quantity"]          = pd.to_numeric(bom["quantity"])

# ── 2. 재고 가용 현황 (stock_ro) ──────────────────────────────
stock_agg = (
    stock.groupby("sku_id")
    .agg(물리재고=("physical_quantity","sum"),
         가용재고=("available_quantity","sum"),
         예약재고=("reserved_quantity","sum"))
    .reset_index()
)

# ── 3. 출고 이행률 (outgoing_item_ro) ─────────────────────────
ogi_og = ogi.merge(og[["id","status"]], left_on="outgoing_id", right_on="id", how="left")
outgoing_agg = (
    ogi_og.groupby("sku_id")
    .agg(출고지시수량=("quantity","sum"),
         실출고수량=("outgoing_quantity","sum"))
    .reset_index()
)
outgoing_agg["미처리수량"] = outgoing_agg["출고지시수량"] - outgoing_agg["실출고수량"]
outgoing_agg["이행률(%)"] = (
    outgoing_agg["실출고수량"] / outgoing_agg["출고지시수량"] * 100
).round(1).fillna(0)

# 미처리 상태별 건수
status_agg = (
    ogi_og[ogi_og["outgoing_quantity"] == 0]
    .groupby("sku_id")["status"]
    .apply(lambda x: " / ".join(sorted(x.value_counts().index)))
    .reset_index()
    .rename(columns={"status":"미처리사유"})
)

# ── 4. 주문/배송 (order_line_ro, delivery_item_ro) ────────────

# 주문수량 + 주문금액 (order_line_ro 기준)
order_agg = (
    ol.groupby("sku_id")
    .agg(주문수량=("quantity","sum"),
         주문금액=("base_currency_amount","sum"))
    .reset_index()
)
order_agg["주문금액"] = order_agg["주문금액"].round(0).astype(int)

# 배송완료수량 (delivery_ro.status='COMPLETED' 기준)
dl_done = dl[dl["status"] == "COMPLETED"][["id"]].rename(columns={"id":"delivery_id"})
dli_done = dli.merge(dl_done, on="delivery_id", how="inner")
delivery_agg = (
    dli_done.groupby("sku_id")["quantity"]
    .sum()
    .reset_index()
    .rename(columns={"quantity":"배송완료수량"})
)

# ── 5. BOM 전개 소요수량 (stock_usage_ro OUTGOING_COMPLETED 기준) ─────────────
# 완제품 레벨 재고 이동 없이 구성요소 레벨에서만 차감 발생하는 구조
bom_finished  = set(bom["sku_id"])
bom_component = set(bom["bundled_sku_id"])

su["delta"] = pd.to_numeric(su["delta"])
bom_req = (
    su[(su["sku_id"].isin(bom_component)) & (su["type"] == "OUTGOING_COMPLETED")]
    .groupby("sku_id")["delta"]
    .sum()
    .abs()
    .reset_index()
    .rename(columns={"delta": "BOM전개소요수량"})
)
bom_req["BOM전개소요수량"] = bom_req["BOM전개소요수량"].astype(int)

# ── 6. 기간 재고 변동 (stock_usage_ro) ───────────────────────
def sum_type(df, t, col):
    return (df[df["type"]==t].groupby("sku_id")["delta"].sum()
            .reset_index().rename(columns={"delta":col}))

incoming_agg   = sum_type(su, "INCOMING_COMPLETED",   "기간_입고")
outgoing_c_agg = sum_type(su, "OUTGOING_COMPLETED",   "기간_출고완료")
adjust_agg     = sum_type(su, "ADJUSTMENT_COMPLETED", "기간_조정")

# ── 7. SKU 마스터 ──────────────────────────────────────────────
def classify_bom(sku_id):
    if sku_id in bom_finished:
        return "BOM완제품"
    elif sku_id in bom_component:
        return "BOM구성요소"
    return "SINGLE"

sku_info = (
    sku[sku["deleted_at"].isna()][["id","name","sku_code","composition_type","sku_group_id"]]
    .rename(columns={"id":"sku_id","name":"sku_nm"})
    .merge(sku_grp[sku_grp["deleted_at"].isna()][["id","biz_partner_id"]],
           left_on="sku_group_id", right_on="id", how="left")
    .drop(columns=["id","sku_group_id"])
)
sku_info["BOM유형"] = sku_info["sku_id"].apply(classify_bom)

# ── 8. 통합 조인 ──────────────────────────────────────────────
base = sku_info.copy()
for df in [stock_agg, outgoing_agg, status_agg,
           incoming_agg, outgoing_c_agg, adjust_agg,
           order_agg, delivery_agg, bom_req]:
    base = base.merge(df, on="sku_id", how="left")

num_cols = ["물리재고","가용재고","예약재고",
            "출고지시수량","실출고수량","미처리수량",
            "기간_입고","기간_출고완료","기간_조정",
            "주문수량","주문금액","배송완료수량","BOM전개소요수량"]
base[num_cols] = base[num_cols].fillna(0).astype(int)
base["이행률(%)"] = base["이행률(%)"].fillna(0)
base["미처리사유"] = base["미처리사유"].fillna("")

# 위험 플래그
def flag(r):
    if r["가용재고"] < r["미처리수량"] and r["미처리수량"] > 0:
        return "⚠️ 재고부족"
    elif r["미처리수량"] > 0:
        return "🔴 미처리"
    elif r["출고지시수량"] > 0:
        return "✅ 이행완료"
    return ""

base["상태"] = base.apply(flag, axis=1)

COLS = [
    "sku_id","sku_nm","sku_code","biz_partner_id","composition_type","BOM유형","상태",
    "물리재고","가용재고","예약재고",
    "출고지시수량","실출고수량","미처리수량","이행률(%)","미처리사유",
    "기간_입고","기간_출고완료","기간_조정",
    "주문수량","주문금액","배송완료수량","BOM전개소요수량"
]
base = base[COLS].sort_values(["biz_partner_id","미처리수량"], ascending=[True,False])

전체     = base.copy()
미처리    = base[base["미처리수량"] > 0].copy()
재고부족  = base[base["상태"] == "⚠️ 재고부족"].copy()
이행완료  = base[base["상태"] == "✅ 이행완료"].copy()

# ── 7. Excel 출력 ─────────────────────────────────────────────
COL_HEADERS = {
    "sku_id":       ("SKU ID",    16),
    "sku_nm":       ("SKU명",     42),
    "sku_code":     ("SKU코드",   18),
    "biz_partner_id":("거래처ID", 18),
    "composition_type":("구성유형",12),
    "BOM유형":      ("BOM유형",    14),
    "상태":         ("상태",       12),
    "물리재고":     ("물리재고",   10),
    "가용재고":     ("가용재고",   10),
    "예약재고":     ("예약재고",   10),
    "출고지시수량": ("출고지시",   10),
    "실출고수량":   ("실출고",     10),
    "미처리수량":   ("미처리",     10),
    "이행률(%)":   ("이행률(%)",  10),
    "미처리사유":   ("미처리사유", 20),
    "기간_입고":   ("기간입고",   10),
    "기간_출고완료":("기간출고",  10),
    "기간_조정":   ("기간조정",   10),
    "주문수량":    ("주문수량",   10),
    "주문금액":    ("주문금액(₩)", 14),
    "배송완료수량":("배송완료",   10),
    "BOM전개소요수량":("BOM소요",  12),
}

NUM_COLS_IDX = {8,9,10,11,12,13,16,17,18,19,20,21,22}  # 1-based (H~N, P~V)

# 헤더 스타일 색상
HEADER_COLORS = {
    "재고가용":  "1F4E79",  # 진파랑  (물리/가용/예약)
    "출고이행":  "833C00",  # 진갈색  (지시/실출고/미처리/이행률/사유)
    "재고변동":  "375623",  # 진녹색  (입고/출고완료/조정)
    "주문배송":  "4B2D8B",  # 진보라  (주문수량/금액/배송완료)
    "기본":      "2F5496",  # 기본파랑
}
SECTION_MAP = {
    "물리재고":"재고가용","가용재고":"재고가용","예약재고":"재고가용",
    "출고지시수량":"출고이행","실출고수량":"출고이행","미처리수량":"출고이행",
    "이행률(%)":"출고이행","미처리사유":"출고이행",
    "기간_입고":"재고변동","기간_출고완료":"재고변동","기간_조정":"재고변동",
    "주문수량":"주문배송","주문금액":"주문배송","배송완료수량":"주문배송","BOM전개소요수량":"주문배송",
}

ROW_FILLS = {
    "⚠️ 재고부족": "FFC7CE",
    "🔴 미처리":   "FFEB9C",
    "✅ 이행완료": "C6EFCE",
}

def write_sheet(ws, df, title):
    """데이터프레임을 워크시트에 쓰고 스타일 적용"""
    # 제목행
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    tc = ws.cell(1, 1, title)
    tc.font  = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    tc.fill  = PatternFill("solid", start_color="11111B")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # 헤더행 (2행)
    for ci, col in enumerate(COLS, 1):
        label, width = COL_HEADERS[col]
        section = SECTION_MAP.get(col, "기본")
        color   = HEADER_COLORS[section]
        cell = ws.cell(2, ci, label)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", start_color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 20

    # 데이터행
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = "#,##0"

    for ri, (_, row) in enumerate(df.iterrows(), 3):
        상태값 = str(row.get("상태",""))
        fill_color = ROW_FILLS.get(상태값)
        for ci, col in enumerate(COLS, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws.cell(ri, ci, val)
            cell.font   = Font(name="Arial", size=9)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if ci in NUM_COLS_IDX and isinstance(val, (int, float)) and col != "이행률(%)":
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if ci == 14:  # 이행률
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if fill_color:
                cell.fill = PatternFill("solid", start_color=fill_color)
    ws.freeze_panes = "H3"

SHEETS = [
    ("전체현황",    전체,    "SKU 통합 현황 (재고가용 + 출고이행 + 기간변동)"),
    ("미처리출고",  미처리,  "미처리 출고 현황 — 물류팀 확인"),
    ("재고부족위험",재고부족,"가용재고 < 미처리수량 — 긴급 발주/조정 필요"),
    ("이행완료",    이행완료,"이행률 100% SKU — 정상 처리 확인"),
]

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    # 빈 시트용 더미 write
    for sname, df, title in SHEETS:
        df.to_excel(writer, sheet_name=sname, index=False, startrow=2)
        ws = writer.sheets[sname]
        # 열 데이터를 실제 헤더/스타일로 교체
        for col_letter in [get_column_letter(i) for i in range(1, len(COLS)+1)]:
            ws[f"{col_letter}3"] = None  # 더미 헤더 제거
        write_sheet(ws, df, title)

bom_counts = base["BOM유형"].value_counts()
print(f"✅ 저장 완료: {OUTPUT_FILE}")
print(f"\n[요약]")
print(f"  전체 SKU:        {len(전체):>5}개")
print(f"  미처리 출고:      {len(미처리):>5}개 SKU")
print(f"  ⚠️  재고부족 위험: {len(재고부족):>5}개 SKU")
print(f"  ✅ 이행 완료:     {len(이행완료):>5}개 SKU")
print(f"\n[BOM 유형]")
for k, v in bom_counts.items():
    print(f"  {k}: {v}개")
